"""Hardening tests for the D3 iframe chart builders + Excel export.

Builds each chart's standalone HTML document and inspects the string directly:

* NaN/Inf in the input records must serialize as ``null`` (never a bare ``NaN``
  token, which is invalid JSON but a valid JS literal that used to plot junk);
* hostile titles / labels must not be able to terminate the inline ``<script>``
  or HTML context;
* formula-like Excel string cells must round-trip as text, not live formulas;
* empty data must not crash the builders, and the y-domain guards must be
  present in the emitted script.

Fast and dependency-light: no browser, no DB — just string parsing plus
``openpyxl.load_workbook`` for the Excel round-trip.
"""
from __future__ import annotations

import json
import re
from io import BytesIO

import pandas as pd
import pytest
from openpyxl import load_workbook

from repower.dashboard.components.excel_export import build_excel_workbook
from repower.dashboard.components.generation_chart import build_generation_chart_html
from repower.dashboard.components.price_chart import build_price_chart_html
from repower.dashboard.components.tieline_chart import build_tieline_chart_html
from repower.dashboard.components.volume_chart import build_volume_chart_html

NAN = float("nan")
INF = float("inf")
NASTY = "</script><script>alert(1)</script>"

PRICE_METRICS = ["price_max", "price_avg", "price_min"]
VOLUME_METRICS = ["demand_mw", "contracted_mw", "missing_mw"]
TIELINE_METRICS = ["upper_limit_fwd", "upper_limit_rev", "reserved_fwd", "reserved_rev"]
STACK_KEYS = ["nuclear", "lng"]

_ALL_KEYS = PRICE_METRICS + VOLUME_METRICS + TIELINE_METRICS + STACK_KEYS + ["area_demand_mw"]
COLORS = {k: "#123456" for k in _ALL_KEYS}
LABELS = {k: k for k in _ALL_KEYS}


def _price(data, title="Tokyo", labels=LABELS):
    return build_price_chart_html(data, PRICE_METRICS, COLORS, labels, title=title)


def _volume(data, title="Tokyo", labels=LABELS):
    return build_volume_chart_html(data, VOLUME_METRICS, COLORS, labels, title=title)


def _tieline(data, title="Hokkaido->Tohoku", labels=LABELS):
    return build_tieline_chart_html(data, TIELINE_METRICS, COLORS, labels, title=title)


def _generation(data, title="Tokyo", labels=LABELS):
    return build_generation_chart_html(data, STACK_KEYS, COLORS, labels, title=title)


def _row(dt: str, keys: list[str], value) -> dict:
    return {"datetime": dt, **{k: value for k in keys}}


# (name, builder, metric keys carried by the data records, y-guard snippet)
CHARTS = [
    ("price", _price, PRICE_METRICS, "yMax = 10;"),
    ("volume", _volume, VOLUME_METRICS, "maxMW = 100;"),
    ("tieline", _tieline, TIELINE_METRICS, "maxVal = 100;"),
    ("generation", _generation, STACK_KEYS + ["area_demand_mw"], "maxY = 100;"),
]


def _raw_data_payload(html: str) -> str:
    m = re.search(r"const rawData = (\[.*?\]);", html, re.S)
    assert m, "rawData payload not found in chart HTML"
    return m.group(1)


def _fail_on_constant(name):  # pragma: no cover - only fires on regression
    raise AssertionError(f"bare {name} token in emitted JSON")


# ── (a) NaN / Inf sanitation ────────────────────────────────────────────────

@pytest.mark.parametrize("name,builder,keys,_guard", CHARTS, ids=[c[0] for c in CHARTS])
def test_nan_serialized_as_null(name, builder, keys, _guard):
    data = [
        _row("2026-07-01T00:00:00", keys, NAN),
        _row("2026-07-01T00:30:00", keys, INF),
        _row("2026-07-01T01:00:00", keys, 5.5),
    ]
    html = builder(data)
    payload = _raw_data_payload(html)

    assert "NaN" not in payload
    assert "Infinity" not in payload

    # Strict parse: json.loads silently accepts NaN/Infinity unless told not to.
    parsed = json.loads(payload, parse_constant=_fail_on_constant)
    for k in keys:
        assert parsed[0][k] is None, f"{name}: NaN {k} should serialize as null"
        assert parsed[1][k] is None, f"{name}: Inf {k} should serialize as null"
        assert parsed[2][k] == 5.5


# ── (b) Injection hardening ─────────────────────────────────────────────────

@pytest.mark.parametrize("name,builder,keys,_guard", CHARTS, ids=[c[0] for c in CHARTS])
def test_hostile_title_and_labels_cannot_break_script_context(name, builder, keys, _guard):
    data = [_row("2026-07-01T00:00:00", keys, 1.0)]
    labels = dict(LABELS)
    labels[keys[0]] = NASTY  # hostile label travels through the JSON payload
    html = builder(data, title=NASTY, labels=labels)

    # The raw sequence never appears: the title is HTML-escaped and the label
    # payload has "</" escaped inside its JSON string.
    assert NASTY not in html
    assert "&lt;/script&gt;" in html  # escaped title still displayed
    # Exactly the two legitimate closers: the d3 <script src> tag + inline script.
    assert html.count("</script>") == 2


# ── (c) Excel formula-injection guard ───────────────────────────────────────

def test_excel_formula_strings_round_trip_as_text():
    df = pd.DataFrame(
        {
            "name": ["=cmd|' /C calc'!A0", "+SUM(A1:A9)", "-2+3", "@foo", "safe"],
            "value": [1, 2, -3, 4.5, None],
        }
    )
    blob = build_excel_workbook({"data": df})
    ws = load_workbook(BytesIO(blob))["data"]

    # Rows 2..5 are formula-leading strings: guarded with a quote, stored as text.
    for i, original in enumerate(df["name"][:4], start=2):
        cell = ws.cell(row=i, column=1)
        assert cell.data_type != "f", f"row {i} became a live formula"
        assert cell.value == "'" + original

    # Plain strings and numbers pass through untouched.
    assert ws.cell(row=6, column=1).value == "safe"
    assert ws.cell(row=2, column=2).value == 1
    assert ws.cell(row=4, column=2).value == -3
    assert ws.cell(row=4, column=2).data_type == "n"


# ── (d) Empty-data y-domain guard smoke ─────────────────────────────────────

@pytest.mark.parametrize("name,builder,keys,guard", CHARTS, ids=[c[0] for c in CHARTS])
def test_empty_data_builds_and_carries_y_guard(name, builder, keys, guard):
    html = builder([])
    assert json.loads(_raw_data_payload(html), parse_constant=_fail_on_constant) == []
    # The non-finite/zero y-max fallback must be present in the script.
    assert guard in html, f"{name}: y-domain guard missing"
    assert "Number.isFinite" in html


# ── Volume band gating ──────────────────────────────────────────────────────

def test_volume_band_is_gated_on_actual_series_presence():
    html = _volume([_row("2026-07-01T00:00:00", VOLUME_METRICS, 1.0)])
    # Band only draws when the records really carry both series, and skips
    # blocks where either endpoint is missing.
    assert "hasBandData" in html
    assert (
        ".defined(d => Number.isFinite(d.demand_mw) && Number.isFinite(d.contracted_mw))"
        in html
    )
