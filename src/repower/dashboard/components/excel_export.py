"""
Build an Excel workbook from a set of named DataFrames.

Generic and dependency-light: one worksheet per dict key, with a header
row taken from the DataFrame columns followed by the data rows.
"""
from __future__ import annotations

from io import BytesIO

import pandas as pd
from openpyxl import Workbook


def _guard_formula(value):
    """Neutralize spreadsheet formula injection for string cells.

    openpyxl turns any string starting with ``=`` into a live formula (and
    ``+``/``-``/``@`` are formula-leading in Excel/other tools), so a hostile
    value like ``=cmd|' /C calc'!A0`` would execute on open. Prefix such
    strings with a single quote so they round-trip as text.
    """
    if isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
        return "'" + value
    return value


def _sanitize_sheet_title(name: str, used: set[str]) -> str:
    """Return an Excel-legal, unique worksheet title.

    Excel sheet titles are limited to 31 characters and may not contain
    any of ``[]:*?/\\``. Duplicate titles are disambiguated with a suffix.
    """
    title = str(name) if name else "Sheet"
    for ch in "[]:*?/\\":
        title = title.replace(ch, "_")
    title = title[:31] or "Sheet"

    base = title
    i = 1
    while title.lower() in used:
        suffix = f"_{i}"
        title = base[: 31 - len(suffix)] + suffix
        i += 1
    used.add(title.lower())
    return title


def build_excel_workbook(sheets: dict[str, pd.DataFrame]) -> bytes:
    """
    Build an ``.xlsx`` workbook with one worksheet per ``sheets`` entry.

    Parameters
    ----------
    sheets : dict mapping sheet name → DataFrame
        Each DataFrame is written with its column names as the header row,
        followed by one row per record. Insertion order is preserved.

    Returns
    -------
    bytes : The workbook file contents.
    """
    wb = Workbook()
    # Remove the default sheet created by openpyxl; we add our own.
    wb.remove(wb.active)

    used_titles: set[str] = set()

    if not sheets:
        wb.create_sheet(title="Sheet1")

    for name, df in sheets.items():
        ws = wb.create_sheet(title=_sanitize_sheet_title(name, used_titles))
        if df is None:
            continue
        # Header row
        ws.append([_guard_formula(str(c)) for c in df.columns])
        # Data rows
        for row in df.itertuples(index=False, name=None):
            ws.append([_guard_formula(v) for v in row])

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
